# extract_images.py - ไฟล์สำหรับฟังก์ชันแยกรูปภาพจาก Excel

import openpyxl
from openpyxl_image_loader import SheetImageLoader
import io
from PIL import Image

def extract_images_from_xlsx(file_path):
    """
    ฟังก์ชันสำหรับแยกรูปภาพทั้งหมดจากไฟล์ Excel (.xlsx)
    
    Args:
        file_path (str): พาธของไฟล์ .xlsx
        
    Returns:
        list: รายการรูปภาพพร้อมข้อมูลประกอบ เช่น sheet_name, cell_address และ image_data
    """
    try:
        # เปิดไฟล์ Excel
        workbook = openpyxl.load_workbook(file_path)
        
        result = []
        
        # วนลูปผ่านทุก sheet ใน workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            image_loader = SheetImageLoader(sheet)
            
            # หาเซลล์ที่มีรูปภาพ
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell_address = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    if image_loader.image_in(cell_address):
                        try:
                            # ดึงรูปภาพจากเซลล์
                            image = image_loader.get(cell_address)
                            
                            # แปลงรูปภาพเป็น bytes สำหรับส่งกลับ
                            img_byte_arr = io.BytesIO()
                            image_format = image.format if image.format else 'PNG'
                            image.save(img_byte_arr, format=image_format)
                            img_byte_arr.seek(0)
                            
                            # เพิ่มข้อมูลรูปภาพลงในผลลัพธ์
                            result.append({
                                'sheet_name': sheet_name,
                                'cell_address': cell_address,
                                'format': image_format,
                                'image_data': img_byte_arr
                            })
                        except Exception as e:
                            print(f"ไม่สามารถดึงรูปภาพจากเซลล์ {cell_address} ใน sheet '{sheet_name}': {str(e)}")
        
        return result
    
    except Exception as e:
        raise Exception(f"เกิดข้อผิดพลาดในการแยกรูปภาพ: {str(e)}")