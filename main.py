from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from PIL import Image as PILImage
import os, shutil, uuid, io

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# 🔒 ใช้ Service Account จาก Secret File ของ Render
SERVICE_ACCOUNT_FILE = '/etc/secrets/n8ncheckslip-7a8a20abc6b5.json'
FOLDER_ID = '1g1UedBOYIcIiqOn2MJxgWnQz4Gk2isTY'  # โฟลเดอร์ Google Drive ปลายทาง

credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE,
    scopes=['https://www.googleapis.com/auth/drive']
)
drive_service = build('drive', 'v3', credentials=credentials)

app = FastAPI()


@app.get("/")
def root():
    return {"message": "API is working"}


@app.post("/extract-images")
async def extract_images(file: UploadFile = File(...)):
    uploaded_links = []
    sheet_image_counts = {}

    try:
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)

        temp_file_path = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        wb = load_workbook(temp_file_path)

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            images = getattr(ws, "_images", [])
            print("Sheet:", ws.title, "Images found:", len(images))
            sheet_image_counts[sheetname] = len(images)

            for img in images:
                try:
                    img_bytes = io.BytesIO(img._data())  # ✅ เคล็ดลับที่ทำให้ดึงได้แม้เป็นภาพแปลก
                    pil_img = PILImage.open(img_bytes)

                    file_name = f"slip_{uuid.uuid4().hex}.jpg"
                    local_path = os.path.join(temp_dir, file_name)
                    pil_img.save(local_path, format="JPEG")

                    media = MediaFileUpload(local_path, mimetype='image/jpeg')
                    file_metadata = {'name': file_name, 'parents': [FOLDER_ID]}
                    uploaded_file = drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id,webViewLink'
                    ).execute()

                    uploaded_links.append(uploaded_file['webViewLink'])

                except Exception as e:
                    print(f"❌ Error processing image: {e}")

        return JSONResponse(content={
            "uploaded_slips": uploaded_links,
            "sheet_image_counts": sheet_image_counts
        })

    except Exception as e:
        print(f"❌ Top-level error: {e}")
        return JSONResponse(content={
            "uploaded_slips": uploaded_links,
            "error": str(e)
        })
